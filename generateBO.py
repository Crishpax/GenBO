#! python
# generateBO.py - generates a back-order report for the day using data pulled from SAP through the GUI scripting API.
# As input it takes: a file containing either all E materials or all PSP for the project, as well as SpMs ending in '*', a file assigning all E mats to areas,
# target folders for output files (main on the network drive, additional on the user's drive)
# and (optional) a date range - default is 2 calendar weeks from now

#====================================================== USING ======================================================


import os, openpyxl, win32com.client, datetime, pyperclip, time, threading, queue, pythoncom, re
from userColData import userColDict         # Contains column name references for zpp_mpl
from cooisColData import cooisColDict       # Contains column name references for coois
from setHolidays import setHolidays           # A list of holiday date strings in the format YYYYMMDD
from openpyxl.utils.cell import get_column_letter
from log import logErrors

#============================================== FUNCTION DEFINITIONS ==============================================

@logErrors
def getBOInput(filepath, sheetName=None):

    # Takes a filepath for an .xlsx containing material(or PSP) numbers in column A and spm ending with '*' in column B
    # (optional) Takes a sheet name to use, otherwise uses the active sheet
    # Returns a list of lists: index 0 - materials, index 1 - spm
        
    if not os.path.exists(filepath):
        print('BO input file not found, exiting.')
        return

    wb = openpyxl.load_workbook(filepath)

    if sheetName == None:
        sheet = wb.active
    else:
        try:
            sheet = wb[sheetName]
        except Exception as exc:
            print('Error: %s' %(exc))
            return

    materialsOrPSP = []
    spm = []

    for row in range(1, sheet.max_row+1):
        matVal = sheet.cell(row=row, column=1).value
        spmVal = sheet.cell(row=row, column=2).value
        if matVal not in (None, "", " ",):
            materialsOrPSP.append(matVal)
        if spmVal not in (None, "", " ",):
            try:
                spmVal = spmVal[:spmVal.index('*')]
            except ValueError:
                print("SpM: %s does not end with a '*'. (Col:B, Row: %s)" %(spmVal, row))
            finally:
                spm.append(spmVal)

    materialsOrPSP = list(dict.fromkeys(materialsOrPSP))
    spm = list(dict.fromkeys(spm))
    
    return [materialsOrPSP, spm]

@logErrors
def getMaterialArea(filepath, sheetName=None):

    # Takes an .xlsx file (or a sheet) with materials in column A and areas in column B,
    # returns a dictionary of that data (key=mat, val=area)
        
    if not os.path.exists(filepath):
        print('Material-area reference file not found, exiting.')
        return

    wb = openpyxl.load_workbook(filepath)

    if sheetName == None:
        sheet = wb.active
    else:
        try:
            sheet = wb[sheetName]
        except Exception as exc:
            print('Error: %s' %(exc))
            return

    areaDict = {}
    for row in range(1, sheet.max_row+1):
        matVal = sheet.cell(row=row, column=1).value
        areaVal = sheet.cell(row=row, column=2).value

        if matVal != None and areaVal != None:
            areaDict.setdefault(matVal, areaVal)

        else:
            print('Data missing in %s in row %s.' %(os.path.basename(filepath), row))
            continue
        
    return areaDict

@logErrors
def getSapWnd():

    # Returns a list of currently active sap windows (up to 6)
    # The objects in the returned list are of type GuiMainWindow (see: SAP GUI Scripting API documentation)

    sapGui = win32com.client.GetObject('SAPGUI')
    app = sapGui.GetScriptingEngine
    conn = app.Children(0)
    
    windows = []

    for sess in conn.Children:
        wnd = sess.Children(0)
        if isinstance(wnd, win32com.client.CDispatch):
            windows.append(wnd)

    if not len(windows) > 0:
        print('No active SAP windows fount, exiting.')
        return
    else:
        return windows

@logErrors
def getBOCoois(wnd_id, orderList):
    
    # Grabs the ID of the SAP window passed between threads and runs a coois report,
    # ensuring that the list key is headers and with the user's default view
    # returns a dictionary of orders with material numbers and order statuses
    

    # CoInitialize pythoncom and get window instance from ID
    if isinstance(wnd_id, win32com.client.CDispatch):
        wnd = wnd_id
    else:
        pythoncom.CoInitialize()
        wnd = win32com.client.Dispatch(pythoncom.CoGetInterfaceAndReleaseStream(wnd_id, pythoncom.IID_IDispatch))
        checkWndTime = 0
        while not isinstance(wnd, win32com.client.CDispatch):
            if checkWindTime >= 5:
                break
            checkWindTime += 1
            time.sleep(1)
        
    if not len(orderList) > 0:
        print('No orders passed to coois, exiting.')
        return
    
    if not isinstance(wnd, win32com.client.CDispatch):
        print('The passed SAP window is no longer active, exiting.')
        return

    # Maximize the window used for this transaction so all the components are rendered
    wnd.Maximize()
    
    # Start the transaction and make sure headers are selected
    wnd.Parent.StartTransaction('coois')
    cooisList = wnd.FindById('usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/cmbPPIO_ENTRY_SC1100-PPIO_LISTTYP')
    if not cooisList.Key == 'PPIOH000':
        cooisList.Key = 'PPIOH000'
    wnd.FindById('usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT').Text = '/BO_PL08'

    # Open up the order input and paste the data, run the report
    wnd.FindById('usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/btn%_S_AUFNR_%_APP_%-VALU_PUSH').Press()
    pyperclip.copy('\r\n'.join(orderList))
    wnd.Parent.FindById('wnd[1]/tbar[0]/btn[24]').Press()
    wnd.Parent.FindById('wnd[1]/tbar[0]/btn[8]').Press()
    pyperclip.copy('')
    wnd.FindById('tbar[1]/btn[8]').Press()

    # Get the table grid object and make sure all rows are rendered
    grid = wnd.FindById('usr/cntlCUSTOM/shellcont/shell/shellcont/shell')
    for row in range(0,grid.RowCount, 25):
        grid.SetCurrentCell(row, grid.ColumnOrder[0])
        while row > (grid.RowCount-24) and row < grid.RowCount:
            grid.SetCurrentCell(row, grid.ColumnOrder[0])
            row += 1

    # Create the dictionary and fill it with order data, return it
    orderData = {}
    for row in range(grid.RowCount):
        ordNum = grid.GetCellValue(row, 'AUFNR')
        matNum = grid.GetCellValue(row, 'MATNR')
        ordStat = grid.GetCellValue(row, 'STTXT')
        orderData.setdefault(ordNum, {})
        orderData[ordNum].setdefault('MATNR', matNum)
        orderData[ordNum].setdefault('STATUS', ordStat)

    return orderData
    
@logErrors
def getSAPDateFormat(datetimeObj, sep='.'):
    # A tiny function converting a datetime object into a SAP-friendly string
    # (DD.MM.YYYY)
    yearStr = str(datetimeObj.year)
    monthStr = str(datetimeObj.month)
    if len(monthStr) == 1:
        monthStr = '0'+monthStr
    dayStr = str(datetimeObj.day)
    if len(dayStr) == 1:
        dayStr = '0'+dayStr
    return sep.join([dayStr, monthStr, yearStr])

@logErrors
def getDatetimeFromSAPDate(sapDate):
    # A reverse function to getSAPDateFormat()
    # Takes a date string in the format of DD.MM.YYYY
    # and converts it into a datetime object
    # If the SAP date string is incorrect, returns None

    regex = re.compile(r'(\d\d).(\d\d).(\d\d\d\d)')
    mo = regex.search(sapDate)
    if mo == None:
        return None
    else:
        day = mo.groups()[0]
        month = mo.groups()[1]
        year = mo.groups()[2]
        return datetime.datetime(int(year), int(month), int(day))
    
@logErrors
def calculateFutureBTDate(workdays=10):             # This is how many FUTURE workdays (excluding today) the BO report will contain (default = 10)
    # Calculate the date of X workdays in the future accouting
    # for weekends and BT holidays
    if not isinstance(workdays, int):
        print('Workdays in the future must be an integer.')
        return
    daysPassed = 0
    endDate = datetime.datetime.now()           # Start calculating from today's date
    while True:
        if daysPassed == workdays:
            break
        endDate += datetime.timedelta(days=1)       # Add one day so you get the next day's date
        daysPassed += 1
        checkYear = str(endDate.year)
        checkMonth = str(endDate.month)
        if len(checkMonth) == 1:
            checkMonth = '0' + checkMonth
        checkDay = str(endDate.day)
        if len(checkDay) == 1:
            checkDay = '0' + checkDay
        while ''.join([checkYear, checkMonth, checkDay]) in setHolidays or endDate.weekday() in [5,6]:   # If that date is a holiday or a weekend, find the next workday
            endDate += datetime.timedelta(days=1)
            checkYear = str(endDate.year)
            checkMonth = str(endDate.month)
            if len(checkMonth) == 1:
                checkMonth = '0' + checkMonth
            checkDay = str(endDate.day)
            if len(checkDay) == 1:
                checkDay = '0' + checkDay

    return endDate

@logErrors
def filterBySpm(data, spmList, spmColIndex, deleteEmptySpm=True):

    filteredData = []
    for spm in spmList:
        filteredBatch = [row for row in data if row[spmColIndex].startswith(spm)]
        filteredData += filteredBatch

    if not deleteEmptySpm:
        filteredBatch = [row for row in data if row[spmColIndex] == '']
        filteredData += filteredBatch

    return filteredData

@logErrors
def generateBO(wnd, matList, spmList, byMaterial=True,runCoois=True, startDate=None, endDate=None, viewName='BO_TRAXX', deleteE=False, deleteEmptySpm=True):
    
    # Validate material list
    if not len(matList) > 0:
        print('Empty material list passed, exiting.')
        return

    # Validate SAP window
    if not isinstance(wnd, win32com.client.CDispatch):
        print('The passed SAP window is no longer active, exiting.')
        return
    
    # Validate date input
    if not isinstance(startDate, datetime.datetime):
        startDate = None

    if not isinstance(endDate, datetime.datetime):
        endDate = None
        
    # Calculate endDate if it's empty - works using company calendar
    if endDate == None:
        endDate = calculateFutureBTDate()

    # Maximize the window used for this transaction so all the components are rendered
    wnd.Maximize()
    
    # Start ZPP_MPL and input basic parameters
    wnd.Parent.StartTransaction('zpp_mpl')
    wnd.FindById('usr/ctxtS_WERKS-LOW').Text = 'PL08'
    wnd.FindById('usr/chkP_T310').Selected = True
    wnd.FindById('usr/ctxtS_BDTER-HIGH').Text = getSAPDateFormat(endDate)
    if isinstance(startDate, datetime.datetime):
        wnd.FindById('usr/ctxtS_BDTER-LOW').Text = getSAPDateFormat(startDate)

    # Set view - TODO figure this shit out
    wnd.FindById('usr/ctxtP_LAYOUT').Text = viewName

    # Input material values if byMaterial=True
    if byMaterial:
        wnd.FindById('usr/btn%_S_MATNRA_%_APP_%-VALU_PUSH').Press()
    # Otherwise input PSP values
    else:
        wnd.FindById('usr/btn%_S_POSIDO_%_APP_%-VALU_PUSH').Press()
    pyperclip.copy('\r\n'.join(matList))
    wnd.Parent.FindById('wnd[1]/tbar[0]/btn[24]').Press()
    wnd.Parent.FindById('wnd[1]/tbar[0]/btn[8]').Press()
    pyperclip.copy('')

    # Run the report
    wnd.FindById('tbar[1]/btn[8]').Press()

    # Get the report grid
    grid = wnd.FindById('usr/cntlGRID1/shellcont/shell')

    # Iterate through all rows to render them
    for row in range(0,grid.RowCount, 25):
        grid.SetCurrentCell(row, grid.ColumnOrder[0])
        while row > (grid.RowCount-24) and row < grid.RowCount:
            grid.SetCurrentCell(row, grid.ColumnOrder[0])
            row += 1

    # Save SAP column names and cell data in variables
    colNames = [col for col in grid.ColumnOrder]                
    cellData = []
    for row in range(grid.RowCount):                                                                
        rowData = []                                            
        for col in grid.ColumnOrder:                        
            rowData.append(grid.GetCellValue(row,col).strip())
        cellData.append(rowData)                        

    # Determine rows to be deleted if they don't match any preferred user spm...

    if 'SERNR' in colNames:
        spmColNum = colNames.index('SERNR')
        cellData = filterBySpm(cellData, spmList, spmColNum, deleteEmptySpm)
        
    # ... or the order number is incorrect...
    
    if 'AUFNR' in colNames:
        ordColNum = colNames.index('AUFNR')
        cellData = [row for row in cellData if row[ordColNum].startswith('5')]

    # ... or there are any non-F type materials... (optional)
    if deleteE and 'BESKZ' in colNames:
        typeColNum = colNames.index('BESKZ')
        cellData = [row for row in cellData if row[typeColNum]== 'F']
    
    # If True is passed for runCoois, create a separate thread to run getBOCoois(),
    # passing it the current SAP window and a list of production orders.
    # If the thread is created, it is passed in a list along with its queue object
    # in order to access the return value of getBOCoois in the main thread
    if runCoois == True and 'AUFNR' in colNames:
        prodOrders = []
        for row in cellData:
            prodOrders.append(row[ordColNum])
        prodOrders = list(dict.fromkeys(prodOrders))
        que = queue.Queue()
        wnd_id = pythoncom.CoMarshalInterThreadInterfaceInStream(pythoncom.IID_IDispatch, wnd)
        cooisThread = threading.Thread(target=lambda q, arg1, arg2: q.put(getBOCoois(arg1, arg2)), args=(que, wnd_id, prodOrders))
        cooisThread.start()
        return [colNames, cellData, [cooisThread, que]]
        
    return [colNames, cellData]

@logErrors
def createBOSheet(colNames, cellData):

    if len(colNames) == 0 or len(cellData) == 0:
        print('Insufficient data passed to create the BO worksheet.')
        return

    desiredCols = ['WERKS',             # SAP columns that need to be in the file
    'MATNR',
    'AUFNR',
    'SERNR',
    'PLNBEZ',
    'BDTER',
    'VORNR',
    'PSPEL',
    'BDMNG',
    'VMENG',
    'ENMNG',
    'LGORT',
    'ZZRMC',
    'FEVOR',
    'EKGRP',
    'KZKRI',
    'ZZRTX',
    'ZZREN',
    'ZZDOD',
    'ZZSNO',
    'ZZSPN',
    'BESKZ']
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    today = datetime.datetime.now()
    dateStyle = openpyxl.styles.NamedStyle(name='datetime', number_format='DD/MM/YYYY')
    writtenColumns = 0
    writtenColRef = {}
    for colName in desiredCols:
        if colName not in colNames:
            print("Couldn't find column '%s' in your view, consider adding it to your SAP view. %s" %(userColDict[colName], colName))
            continue
        colIndex = colNames.index(colName)
        # Write the column header
        sheet.cell(row=1, column=writtenColumns+1).value = userColDict[colName]
        # Write column content
        for i in range(len(cellData)):
            targetCell = sheet.cell(row=i+2, column=writtenColumns+1)
            targetData = cellData[i][colIndex]
            # In case of requirement dates, format datetime
            if colName == 'BDTER' or colName == 'ZZDOD':
                if len(targetData) == 10:
                    year = targetData[-4:]
                    month = targetData[3:5]
                    day = targetData[:2]
                    dateVal = datetime.datetime(int(year), int(month), int(day))
                    targetCell.value = dateVal
                    targetCell.style = dateStyle
            # In case of numeric strings, format them as proper floats
            elif len(targetData) > 4 and targetData[-4] == ',' and targetData[-3:].isnumeric():
                digitsBefore = int(targetData[:targetData.index(',')])
                digitsAfter = int(targetData[targetData.index(',')+1:])/1000
                targetCell.value = (digitsBefore+digitsAfter)
                
            else:
                # Try to convert any data into ints
                try:
                    targetCell.value = int(targetData)
                # If not possible, store as string
                except:
                    targetCell.value = targetData
        writtenColRef.setdefault(colName, writtenColumns+1)
        writtenColumns += 1

    # Create values for the key column
    
    if 'MATNR' in writtenColRef and 'AUFNR' in writtenColRef:
        keyVals = []
        for row in range(2, sheet.max_row+1):
            matCell = sheet.cell(row=row, column=writtenColRef['MATNR'])
            ordCell = sheet.cell(row=row, column=writtenColRef['AUFNR'])
            key = str(matCell.value) + str(ordCell.value)
            keyVals.append(key)

    # Determine where the column should be inserted
    
    if 'BDTER' in writtenColRef:
        keyColIndex = writtenColRef['BDTER']
    elif 'PLNBEZ' in writtenColRef:
        keyColIndex = writtenColRef['PLNBEZ'] + 1
    else:
        keyColIndex = 4

    # Insert the column and write the values
    
    sheet.insert_cols(keyColIndex)
    for column in writtenColRef:
        if writtenColRef[column] >= keyColIndex:
            writtenColRef[column] += 1
            
    sheet.cell(row=1, column=keyColIndex).value = 'Key'
    for row in range(2, sheet.max_row+1):
        sheet.cell(row=row, column=keyColIndex).value = keyVals[row-2]
    writtenColRef.setdefault('Key', keyColIndex)
   
    sheet.title = 'BO '+'.'.join([str(today.day), str(today.month), str(today.year)])
    return [wb, writtenColRef]

@logErrors    
def formatBO(wb, writtenColRef, orderDict, areaDict):
    # Validate arguments
    if not isinstance(wb, openpyxl.workbook.workbook.Workbook):
        print('No proper workbook passed to format.')
        return
    if not len(orderDict) > 0:
        print('No order data passed to the program.')
    if not len(areaDict) > 0:
        print('No material-area assignment passed to the program.')
    #if not os.path.exists(targFolder):
    #    print('The specified target folder does not exist.')
    #if userFolder is not '' and not os.path.exists(userFolder):
    #    print('The specified secondary folder does not exist.')

    colWidthData = {'AUFNR': 12.0,      # Width values for each desired column
     'BDMNG': 9.0,
     'BDTER': 12.0,
     'BESKZ': 8,
     'EKGRP': 10.0,
     'ENMNG': 9.5,
     'FEVOR': 10.0,
     'KZKRI': 2,
     'LGORT': 6,
     'MATNR': 18.0,
     'PLNBEZ': 18.0,
     'PSPEL': 21.0,
     'SERNR': 10,
     'VMENG': 9.0,
     'VORNR': 9.0,
     'WERKS': 8.0,
     'ZZRMC': 10.0,
     'ZZSNO': 9.0,
     'ZZSPN': 28.0,
    'ZZRTX': 14,
    'ZZREN': 12,
    'ZZDOD': 12,
    'Key': 30.0,
    'Obszar': 12,
    'Komentarz planowania': 45,
    'Status zlecenia': 45,
    'Komentarz ze spotkania':40,
    'Komentarz MRP':40,
    'Komentarz zakupy':40,
    'Data zakupy':12.0,
    'Różnica':12.0,
    'Missing parts': 9.0}

    # Handle the material column - add it if it's not there from SAP
    
    sheet = wb.active
    
    if 'PLNBEZ' in writtenColRef:
        sapMaterialColumn = True
    else:
        sapMaterialColumn = False
        matColIndex = writtenColRef['SERNR']+1
        sheet.insert_cols(matColIndex)
        sheet.cell(row=1, column=matColIndex).value = 'Materiał'
        sheet.column_dimensions[get_column_letter(matColIndex)].width = 18
        for column in writtenColRef:
            if writtenColRef[column] >= matColIndex:
                writtenColRef[column] += 1
        writtenColRef.setdefault('Materiał', matColIndex)

    # Add standard user columns after 'Key'
    #standardUserCols = ['Obszar', 'Komentarz planowania', 'Status zlecenia', 'Komentarz ze spotkania', 'Komentarz MRP', 'Komentarz zakupy', 'Data zakupy', 'Różnica']
    standardUserCols = ['Różnica', 'Data zakupy', 'Komentarz zakupy', 'Komentarz MRP', 'Komentarz ze spotkania', 'Status zlecenia', 'Komentarz planowania', 'Obszar']
    for col in standardUserCols:
        colIndex = writtenColRef['Key'] + 1
        sheet.insert_cols(colIndex)
        sheet.cell(row=1, column=colIndex).value = col
        for column in writtenColRef:
            if writtenColRef[column] >= colIndex:
                writtenColRef[column] += 1
        writtenColRef.setdefault(col, colIndex)

    # If order numbers are available, fill order status values
    # If order material was not prefilled by SAP, fill order material values
    if 'AUFNR' in writtenColRef and len(orderDict) >0:
        for row in range(2, sheet.max_row+1):
            ordNum = str(sheet.cell(row=row, column=writtenColRef['AUFNR']).value)
            statCell = sheet.cell(row=row, column=writtenColRef['Status zlecenia'])
            matCell = sheet.cell(row=row, column=writtenColRef['Materiał'])
            try:
                statCell.value = orderDict[ordNum]['STATUS']
            except KeyError:
                print('Could not find order %s in the coois report.' %(ordNum))
            if sapMaterialColumn == False:
                try:
                    matCell.value = orderDict[ordNum]['MATNR']
                except KeyError:
                    continue

    # Fill in area data from areaDict
    for row in range(2, sheet.max_row+1):
        matNum = str(sheet.cell(row=row, column=writtenColRef['Materiał']).value)
        areaCell = sheet.cell(row=row, column=writtenColRef['Obszar'])
        try:
            areaCell.value = areaDict[matNum]
        except:
            print('Could not find area assigned to material %s.' %(matNum))
            areaCell.value = 'BRAK'

    # Format 'Data zakupy' to datetime (without inserting any values)
    dateStyle = openpyxl.styles.NamedStyle(name='datetime2', number_format='DD/MM/YYYY')
    for row in range(2, sheet.max_row+1):
        dateCell = sheet.cell(row=row, column=writtenColRef['Data zakupy'])
        dateCell.style = dateStyle

    # Insert formulas into 'Różnica'
    reqColLetter = get_column_letter(writtenColRef['BDTER'])
    delColLetter = get_column_letter(writtenColRef['Data zakupy'])
    for row in range(2, sheet.max_row+1):
        diffCell = sheet.cell(row=row, column=writtenColRef['Różnica'])
        diffCell.value = '='+reqColLetter+str(row)+'-'+delColLetter+str(row)

    # Insert the 'Missing parts' column and calculate values for it
    if 'BDMNG' in writtenColRef and 'VMENG' in writtenColRef and 'ENMNG' in writtenColRef:
        missingColIndex = writtenColRef['ENMNG']+1
        sheet.insert_cols(missingColIndex)
        for column in writtenColRef:
            if writtenColRef[column] >= missingColIndex:
                writtenColRef[column] += 1
        sheet.cell(row=1, column=missingColIndex).value = 'Missing parts'
        writtenColRef.setdefault('Missing parts', missingColIndex)
        for row in range(2, sheet.max_row+1):
            try:
                reqAmount = float(sheet.cell(row=row, column=writtenColRef['BDMNG']).value)
                confAmount = float(sheet.cell(row=row, column=writtenColRef['VMENG']).value)
                delAmount = float(sheet.cell(row=row, column=writtenColRef['ENMNG']).value)
                mpCell = sheet.cell(row=row, column=writtenColRef['Missing parts'])
                mpCell.value = float(reqAmount-confAmount-delAmount)
            except:
                print(str(row) + ' problem with MP')
                continue
    
    # Apply column width
    for column in writtenColRef:
        if column in colWidthData:
            colLetter = get_column_letter(writtenColRef[column])
            sheet.column_dimensions[colLetter].width = colWidthData[column]

    # Apply header styling
    headerSide = openpyxl.styles.Side(border_style='thin', color='000000')
    headerBorder = openpyxl.styles.Border(left = headerSide, right = headerSide, top = headerSide, bottom = headerSide)
    headerFont = openpyxl.styles.Font(bold = True)
    headerAlignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    headerFill = openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='FFF2CC'))
    headerStyle = openpyxl.styles.NamedStyle(name='boheader', font = headerFont, fill = headerFill, border = headerBorder, alignment = headerAlignment)
    for col in range(1, sheet.max_column+1):
        sheet.cell(row=1, column=col).style = headerStyle

    return[wb,writtenColRef]

@logErrors
def findLatestBOInDirectory(directory):
    
    # Scans a folder for the latest BO file,
    # returs [0] the latest BO file path, [1] a date match object
    # where groups [0][2][4] contain year, month and day
    # and groups [1][3] contain optional separators
    
    if not os.path.exists(directory):
        print('Target directory not found.')
        return

    print('in find latest')
    
    boDateRegex = re.compile(r'(\d\d\d\d)(.|-)?(\d\d)(.|-)?(\d\d)(_\d\d\d\d_)?') 
    dateDict={}
    for fileName in os.listdir(directory):
        if fileName.endswith('.xlsx') and not fileName.startswith('~$'):
            mo = boDateRegex.search(fileName)
            if mo is not None:
                year = mo.groups()[0]
                month = mo.groups()[2]
                day = mo.groups()[4]
                if int(month) < 13 and int(day) < 32:
                    thisDate = datetime.datetime(int(year), int(month), int(day))
                    if thisDate not in dateDict:
                        dateDict.setdefault(thisDate, [fileName[:-5]])
                    else:
                        dateDict[thisDate].append(fileName[:-5])
                else:
                    continue
    latestDate = max(dateDict)
    latestBO = None
    if len(dateDict[latestDate]) == 1:
        latestBO = dateDict[latestDate][0] + '.xlsx'
    else:
        modTimeDict = {}
        for fileName in dateDict[latestDate]:
            modTime = os.path.getmtime(os.path.join(directory, (fileName+'.xlsx')))
            modTimeDict.setdefault(modTime, fileName)
        latestBO = modTimeDict[max(modTimeDict)] + '.xlsx'

    if latestBO == None:
        return
    else:
        mo = boDateRegex.search(latestBO)
        
        return[os.path.join(directory, latestBO), mo]

@logErrors    
def getNewBOFilename(latestBOPath, mo):
    # Takes a path to the latest BO file and a match object for date in the filename
    # Returns a new filename for today's report
    dateStartIndex = mo.start()
    dateEndIndex = mo.end()
    dateSeparator = mo.groups()[1]
    baseBOName = os.path.basename(latestBOPath)
    beforeDate = baseBOName[:dateStartIndex]
    afterDate = baseBOName[dateEndIndex:]
    today = datetime.datetime.now()
    if dateSeparator == None:
        dateSeparator=''
    yearStr = str(today.year)
    monthStr = str(today.month)
    dayStr = str(today.day)
    if len(monthStr) == 1:
        monthStr = '0'+monthStr
    if len(dayStr) == 1:
        dayStr = '0'+dayStr
    dateString = dateSeparator.join([yearStr, monthStr, dayStr])
    mrpRegex = re.compile(r'(\s)*kom(\s)*mrp(\s)*', re.I)
    proRegex = re.compile(r'(\s)*zak(upy)?(\s)*', re.I)
    verRegex = re.compile(r'_v(\d)+')
    mrpMo = mrpRegex.search(afterDate)
    proMo = proRegex.search(afterDate)
    verMo = verRegex.search(afterDate)
    for mo in [mrpMo, proMo, verMo]:
        if mo is not None:
            afterDate = afterDate[:mo.start()]+afterDate[mo.end():]
    newBOName = os.path.join(os.path.dirname(latestBOPath),(beforeDate + dateString + afterDate))
    if os.path.exists(newBOName):
        ver = 2
        ogLength = len(newBOName)-5
        while os.path.exists(newBOName):
            newBOName = newBOName[:ogLength] + '_v%s.xlsx' %(ver)
            ver += 1
    return newBOName

@logErrors
def getPlannerComFromLastBO(latestBOPath, deliveryDates = False):
    # Opens the latest BO file (path passed as argument)
    # Returns a dictionary of {line key:planner comment}
    # (optional) also includes delivery dates {line key:[planner comment, delivery date]}
    print(latestBOPath, deliveryDates)
    print('kom 1')
    try:
        wb = openpyxl.load_workbook(latestBOPath)
    except Exception as exc:
        print('Could not open the latest BO file to get planner comments, error: %s' %(exc))
    sheet = None
    for sheetname in wb.sheetnames:
        if 'bo' in sheetname.lower() or 'zpp_mpl' in sheetname.lower():
            sheet = wb[sheetname]
    if sheet == None:
        if 'Sheet1' in wb.sheetnames:
            sheet = wb['Sheet1']
    if sheet == None:
        sheet = wb.active
    print('kom 2')

    # Find indices for planner com and key columns
    keyColIndex = None
    plannerColIndex = None
    for column in range(1, sheet.max_column+1):
        if 'komentarz planowan' in str(sheet.cell(row=1, column=column).value).lower() or 'komentarz - planow' in str(sheet.cell(row=1, column=column).value).lower():
            plannerColIndex = column
        if str(sheet.cell(row=1, column=column).value).lower() == 'klucz' or str(sheet.cell(row=1, column=column).value).lower() == 'key':
            keyColIndex = column
    print('kom 3')
    # Find the index of the delivery date column
    if deliveryDates:
        delColIndex = None
        for column in range(1, sheet.max_column+1):
            if str(sheet.cell(row=1, column=column).value).strip().lower() == 'data zakupy':
                delColIndex = column
        if delColIndex == None:     # If the column is not called 'data zakupy', try 'komentarz zakupy'
            for column in range(1, sheet.max_column+1):
                if str(sheet.cell(row=1, column=column).value).strip().lower() == 'komentarz zakupy':
                    delColIndex = column
    if keyColIndex is None and plannerColIndex is None:
        print('Could not find key and planner comment columns in the latest BO file, exiting.')
        return
    # If key values are Excel formulas, refer to
    # 'nr zlecenia' and 'komponent' values instead
    canUseKey = True
    if keyColIndex is None:
        canUseKey = False
    # Check a sample of key cells for formulas
    if canUseKey == True:
        for row in range(2, sheet.max_row+1, (sheet.max_row+1)//((sheet.max_row+1)//10)):
            if str(sheet.cell(row=row, column=keyColIndex).value).startswith('='):
                canUseKey = False
                break
    if canUseKey == False:
        for column in range(1, sheet.max_column+1):
            if str(sheet.cell(row=1, column=column).value).lower() in ['nr zlecenia', 'nr zlec.']:
                ordColIndex = column
            if str(sheet.cell(row=1, column=column).value).lower() == 'komponent':
                matColIndex = column
    keyCommentDict = {}
    print(canUseKey)
    if canUseKey == True:
        for row in range(2, sheet.max_row+1):
            keyCell = sheet.cell(row=row, column=keyColIndex)
            try:
                comment = sheet.cell(row=row, column=plannerColIndex).value
            except:
                comment = ''
            if deliveryDates and delColIndex != None:
                delCell = sheet.cell(row=row, column = delColIndex)
                keyCommentDict.setdefault(str(keyCell.value), [comment, delCell.value])
            else:
                keyCommentDict.setdefault(str(keyCell.value), comment)
    else:
        for row in range(2, sheet.max_row+1):
            ordCell = sheet.cell(row=row, column=ordColIndex)
            matCell = sheet.cell(row=row, column=matColIndex)
            try:
                comment = sheet.cell(row=row, column=plannerColIndex).value
            except:
                comment = ''
            if deliveryDates and delColIndex != None:
                delCell = sheet.cell(row=row, column = delColIndex)
                keyCommentDict.setdefault(str(matCell.value)+str(ordCell.value), [comment, delCell.value])
            else:
                keyCommentDict.setdefault(str(matCell.value)+str(ordCell.value), comment)
    return keyCommentDict

@logErrors
def finishSaveBO(wb, writtenColRef, targetDir1, deliveryDates=False):
    # Validate paths
    if not os.path.exists(targetDir1):
        print('Invalid target directory.')

    # Get data about the latest BO file in targetDir1, including a planner comment dictionary
    # Also generate the path for the new file
    latestBOData = findLatestBOInDirectory(targetDir1)
    latestBOPath = latestBOData[0]
    latestBOMo = latestBOData[1]
    newBOPath = getNewBOFilename(latestBOPath, latestBOMo)
    newBOName = os.path.basename(newBOPath)
    plannerComDict = getPlannerComFromLastBO(latestBOPath, deliveryDates)
    sheet = wb.active
    if deliveryDates:
        dateStyle = openpyxl.styles.NamedStyle(name='datestyle3', number_format='DD/MM/YYYY')
    # Paste planner comments to specific lines
    for row in range(2, sheet.max_row+1):
        keyCell = sheet.cell(row=row, column=writtenColRef['Key'])
        comCell = sheet.cell(row=row, column=writtenColRef['Komentarz planowania'])
        delCell = sheet.cell(row=row, column=writtenColRef['Data zakupy'])
        
        if str(keyCell.value) in plannerComDict:
            if deliveryDates:
                comCell.value = plannerComDict[str(keyCell.value)][0]
                delCell.value = plannerComDict[str(keyCell.value)][1]
                delCell.style = dateStyle
            else:
                comCell.value = plannerComDict[str(keyCell.value)]
    sheet.auto_filter.ref = sheet.dimensions
    wb.save(newBOPath)          # Save the file to the primary directory
    print('File saved under %s .' %(newBOPath))

    return newBOPath
    
if __name__ == "__main__":
    pass
